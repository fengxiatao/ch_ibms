import json
import re
from pathlib import Path

import openpyxl


DEFAULT_EXCEL_PATH = Path("yudao-ui-admin-vue3/public/IBMS编码规范清单_V2.0.xlsx")


DICT_TYPES = [
    {
        "name": "IBMS 专业分组",
        "type": "ibms_group",
        "remark": "来源：IBMS编码规范清单_V2.0.xlsx / 专业分组定义",
    },
    {
        "name": "IBMS 系统码",
        "type": "ibms_system",
        "remark": "来源：IBMS编码规范清单_V2.0.xlsx / 系统码定义",
    },
    {
        "name": "IBMS 点位类型码",
        "type": "ibms_point_type",
        "remark": "来源：IBMS编码规范清单_V2.0.xlsx / 点位类型码定义",
    },
    {
        "name": "IBMS 设备型号码",
        "type": "ibms_device_model",
        "remark": "来源：IBMS编码规范清单_V2.0.xlsx / 设备型号码定义",
    },
    {
        "name": "IBMS 设备类型码",
        "type": "ibms_device_type",
        "remark": "来源：IBMS编码规范清单_V2.0.xlsx / 设备类型码定义",
    },
    {
        "name": "IBMS 区域码",
        "type": "ibms_region",
        "remark": "来源：IBMS编码规范清单_V2.0.xlsx / 区域码定义",
    },
]


def _to_str(v):
    if v is None:
        return ""
    return str(v).strip()


def _split_csv_codes(s: str):
    s = _to_str(s)
    if not s:
        return []
    # 支持 "VI, AL, GR" / "VI,AL,GR" / "VI；AL" / "BA/EP/EL" 等
    parts = re.split(r"[,，;；/\s]+", s)
    return [p.strip() for p in parts if p.strip()]


def _json_remark(obj):
    if not obj:
        return ""
    return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))


def _sql_str(s: str):
    # MySQL 单引号转义
    return "'" + (s or "").replace("\\", "\\\\").replace("'", "''") + "'"


def _ins_dict_type(name: str, typ: str, remark: str):
    return (
        "INSERT INTO `system_dict_type` (`name`, `type`, `status`, `remark`, `creator`, `create_time`, `updater`, `update_time`, `deleted`, `deleted_time`)\n"
        f"SELECT {_sql_str(name)}, {_sql_str(typ)}, 0, {_sql_str(remark)}, 'system', NOW(), 'system', NOW(), b'0', NULL\n"
        f"WHERE NOT EXISTS (SELECT 1 FROM `system_dict_type` WHERE `type` = {_sql_str(typ)});\n"
    )


def _ins_dict_data(sort: int, label: str, value: str, dict_type: str, color_type: str = "", css_class: str = "", remark: str = ""):
    return (
        "INSERT INTO `system_dict_data` (`sort`, `label`, `value`, `dict_type`, `status`, `color_type`, `css_class`, `remark`, `creator`, `create_time`, `updater`, `update_time`, `deleted`)\n"
        f"SELECT {int(sort)}, {_sql_str(label)}, {_sql_str(value)}, {_sql_str(dict_type)}, 0, {_sql_str(color_type)}, {_sql_str(css_class)}, {_sql_str(remark)}, 'system', NOW(), 'system', NOW(), b'0'\n"
        f"WHERE NOT EXISTS (SELECT 1 FROM `system_dict_data` WHERE `dict_type` = {_sql_str(dict_type)} AND `value` = {_sql_str(value)});\n"
    )

def _batch_insert_dict_type(rows):
    # rows: [{name,type,remark}]
    if not rows:
        return ""
    selects = []
    for r in rows:
        selects.append(
            f"SELECT {_sql_str(r['name'])} AS `name`, {_sql_str(r['type'])} AS `type`, {_sql_str(r.get('remark',''))} AS `remark`"
        )
    derived = "\nUNION ALL\n".join(selects)
    return (
        "INSERT INTO `system_dict_type` (`name`, `type`, `status`, `remark`, `creator`, `create_time`, `updater`, `update_time`, `deleted`, `deleted_time`)\n"
        "SELECT v.`name`, v.`type`, 0, v.`remark`, 'system', NOW(), 'system', NOW(), b'0', NULL\n"
        "FROM (\n"
        + derived
        + "\n) v\n"
        "LEFT JOIN `system_dict_type` t ON t.`type` = v.`type` AND t.`deleted` = b'0'\n"
        "WHERE t.`id` IS NULL;\n"
    )


def _batch_insert_dict_data(rows):
    # rows: [{sort,label,value,dict_type,color_type,css_class,remark}]
    if not rows:
        return ""
    selects = []
    for r in rows:
        selects.append(
            "SELECT "
            f"{int(r['sort'])} AS `sort`, "
            f"{_sql_str(r['label'])} AS `label`, "
            f"{_sql_str(r['value'])} AS `value`, "
            f"{_sql_str(r['dict_type'])} AS `dict_type`, "
            f"{_sql_str(r.get('color_type',''))} AS `color_type`, "
            f"{_sql_str(r.get('css_class',''))} AS `css_class`, "
            f"{_sql_str(r.get('remark',''))} AS `remark`"
        )
    derived = "\nUNION ALL\n".join(selects)
    return (
        "INSERT INTO `system_dict_data` (`sort`, `label`, `value`, `dict_type`, `status`, `color_type`, `css_class`, `remark`, `creator`, `create_time`, `updater`, `update_time`, `deleted`)\n"
        "SELECT v.`sort`, v.`label`, v.`value`, v.`dict_type`, 0, v.`color_type`, v.`css_class`, v.`remark`, 'system', NOW(), 'system', NOW(), b'0'\n"
        "FROM (\n"
        + derived
        + "\n) v\n"
        "LEFT JOIN `system_dict_data` d ON d.`dict_type` = v.`dict_type` AND d.`value` = v.`value` AND d.`deleted` = b'0'\n"
        "WHERE d.`id` IS NULL;\n"
    )


def _read_table(ws, columns, start_row=2):
    """
    columns: dict field_name -> column_letter, like {"code":"A","name":"B"}
    stops when primary code column empty
    """
    code_col = columns.get("code") or columns.get("value") or "A"
    row = start_row
    while True:
        code = _to_str(ws[f"{code_col}{row}"].value)
        if not code:
            break
        data = {}
        for k, col in columns.items():
            data[k] = _to_str(ws[f"{col}{row}"].value)
        yield row, data
        row += 1


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Generate IBMS dict seed SQL from Excel.")
    parser.add_argument(
        "--excel",
        type=str,
        default=str(DEFAULT_EXCEL_PATH),
        help="Excel file path. Default: yudao-ui-admin-vue3/public/IBMS编码规范清单_V2.0.xlsx",
    )
    parser.add_argument("--out", type=str, default="", help="Output .sql file path (UTF-8). If empty, prints to stdout.")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        # 相对路径按项目根目录解析
        excel_path = Path(__file__).resolve().parent.parent / excel_path
    excel_path = excel_path.resolve()

    if not excel_path.exists():
        raise SystemExit(
            "Excel not found: "
            f"{excel_path}\n"
            "请将《IBMS编码规范清单_V2.0.xlsx》放到默认路径，或使用参数指定：\n"
            "  python tools/gen_ibms_dict_seed_sql.py --excel <path-to-xlsx> --out ruoyi-vue-pro/sql/mysql/ibms_dict_seed.sql"
        )

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    out = []
    dict_type_rows = []
    dict_data_rows = []
    out.append("-- =============================================\n")
    out.append("-- IBMS 编码规范字典初始化脚本（MySQL）\n")
    out.append(f"-- 来源：{excel_path.name}\n")
    out.append("-- 可重复执行：已存在时跳过\n")
    out.append("-- =============================================\n\n")

    for dt in DICT_TYPES:
        dict_type_rows.append({"name": dt["name"], "type": dt["type"], "remark": dt["remark"]})

    # 专业分组
    ws = wb["专业分组定义"]
    sort = 1
    for _, r in _read_table(
        ws,
        {
            "code": "A",
            "name": "B",
            "systems": "C",
            "icon": "D",
            "color": "E",
            "desc": "F",
            "systemCount": "G",
        },
    ):
        remark = _json_remark(
            {
                "systems": _split_csv_codes(r["systems"]),
                "icon": r["icon"] or None,
                "color": r["color"] or None,
                "desc": r["desc"] or None,
                "systemCount": int(r["systemCount"]) if r["systemCount"].isdigit() else None,
            }
        )
        dict_data_rows.append(
            {"sort": sort, "label": r["name"], "value": r["code"], "dict_type": "ibms_group", "remark": remark}
        )
        sort += 1

    # 系统码
    ws = wb["系统码定义"]
    sort = 1
    for _, r in _read_table(ws, {"code": "A", "name": "B", "en": "C", "groupName": "D", "desc": "E"}):
        # group 既可能是“智慧安防”也可能是代码，这里优先映射为代码：通过专业分组表反查
        group_code = ""
        group_name = r["groupName"]
        if group_name:
            # 反查专业分组
            for _, gr in _read_table(wb["专业分组定义"], {"code": "A", "name": "B"}):
                if gr["name"] == group_name or gr["code"] == group_name:
                    group_code = gr["code"]
                    break
        remark = _json_remark(
            {
                "group": group_code or group_name or None,
                "en": r["en"] or None,
                "desc": r["desc"] or None,
            }
        )
        dict_data_rows.append(
            {"sort": sort, "label": r["name"], "value": r["code"], "dict_type": "ibms_system", "remark": remark}
        )
        sort += 1

    # 点位类型
    ws = wb["点位类型码定义"]
    sort = 1
    for _, r in _read_table(ws, {"code": "A", "name": "B", "dataType": "C", "systems": "D", "desc": "E"}):
        remark = _json_remark(
            {
                "dataType": r["dataType"] or None,
                "systems": _split_csv_codes(r["systems"]),
                "desc": r["desc"] or None,
            }
        )
        dict_data_rows.append(
            {"sort": sort, "label": r["name"], "value": r["code"], "dict_type": "ibms_point_type", "remark": remark}
        )
        sort += 1

    # 设备型号码
    ws = wb["设备型号码定义"]
    sort = 1
    for _, r in _read_table(ws, {"code": "A", "name": "B", "system": "C", "desc": "D"}):
        remark = _json_remark({"system": r["system"] or None, "desc": r["desc"] or None})
        dict_data_rows.append(
            {
                "sort": sort,
                "label": r["name"],
                "value": r["code"],
                "dict_type": "ibms_device_model",
                "remark": remark,
            }
        )
        sort += 1

    # 设备类型码
    ws = wb["设备类型码定义"]
    sort = 1
    for _, r in _read_table(ws, {"code": "A", "name": "B", "desc": "C"}):
        remark = _json_remark({"desc": r["desc"] or None})
        dict_data_rows.append(
            {"sort": sort, "label": r["name"], "value": r["code"], "dict_type": "ibms_device_type", "remark": remark}
        )
        sort += 1

    # 区域码
    ws = wb["区域码定义"]
    sort = 1
    for _, r in _read_table(ws, {"code": "A", "name": "B", "regionType": "C", "example": "D"}):
        remark = _json_remark(
            {
                "regionType": r["regionType"] or None,
                "example": r["example"] or None,
            }
        )
        dict_data_rows.append(
            {"sort": sort, "label": r["name"], "value": r["code"], "dict_type": "ibms_region", "remark": remark}
        )
        sort += 1

    out.append(_batch_insert_dict_type(dict_type_rows))
    out.append("\n")
    out.append(_batch_insert_dict_data(dict_data_rows))
    out.append("\n")
    sql = "".join(out)

    if args.out:
        out_path = Path(args.out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(sql, encoding="utf-8", newline="\n")
    else:
        print(sql)


if __name__ == "__main__":
    main()

